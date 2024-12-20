import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# workbook = xl.load_workbook("transactions.xlsx")
# sheet = workbook["Sheet1"]
# cell = sheet["a1"]
# cell = sheet.cell(1,1)
# print(cell.value)

# print(sheet.max_column) # total columns in the sheet
# print(sheet.max_row) # total rows in the sheet

def automate_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook["Sheet1"]
    for row in range(2, sheet.max_row + 1): # start from 2 to ignore headings
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        print(cell.value,"->",corrected_price)

        #new cell containing new value
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
    

    # use Reference class to select range of values.
    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,"e2")
    # workbook.save("transactions2.xlsx")
    #override the original file if program is working fine
    workbook.save(filename)
    
    
automate_workbook("transactions.xlsx")